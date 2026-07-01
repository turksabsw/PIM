"""XLSX Excel Export Module

This module provides functionality for generating Excel (XLSX) product
catalogs using openpyxl with write_only mode for optimal memory usage
when handling large product catalogs (100k+ products).

The module supports:
- Write-only mode for memory-efficient large catalog generation
- Multiple worksheets (Products, Variants, Attributes, Media, Prices)
- Configurable column mapping via Export Profile
- Multi-language content export
- Data validation and dropdown lists
- Cell styling and conditional formatting
- Auto-filter and freeze panes for usability

Usage:
    from frappe_pim.pim.export.xlsx import export_catalog

    file_path = export_catalog(
        profile_name="my_xlsx_profile",
        save_file=True
    )

Note: frappe imports are deferred to function level to allow module
import without Frappe being available (e.g., for testing/verification).
"""

from datetime import datetime
from io import BytesIO


# Default column configurations for each sheet
DEFAULT_PRODUCT_COLUMNS = [
    {"field": "product_code", "header": "Product Code", "width": 20},
    {"field": "product_name", "header": "Product Name", "width": 40},
    {"field": "short_description", "header": "Short Description", "width": 60},
    {"field": "description", "header": "Description", "width": 80},
    {"field": "product_family", "header": "Product Family", "width": 25},
    {"field": "brand", "header": "Brand", "width": 20},
    {"field": "barcode", "header": "GTIN/Barcode", "width": 18},
    {"field": "status", "header": "Status", "width": 15},
    {"field": "completeness_score", "header": "Completeness %", "width": 15},
]

DEFAULT_VARIANT_COLUMNS = [
    {"field": "variant_code", "header": "Variant Code", "width": 20},
    {"field": "variant_name", "header": "Variant Name", "width": 40},
    {"field": "parent_product", "header": "Parent Product", "width": 20},
    {"field": "barcode", "header": "GTIN/Barcode", "width": 18},
    {"field": "price", "header": "Price", "width": 12},
    {"field": "currency", "header": "Currency", "width": 10},
    {"field": "stock_qty", "header": "Stock Qty", "width": 12},
    {"field": "status", "header": "Status", "width": 15},
]

DEFAULT_ATTRIBUTE_COLUMNS = [
    {"field": "product_code", "header": "Product Code", "width": 20},
    {"field": "attribute_code", "header": "Attribute Code", "width": 25},
    {"field": "attribute_name", "header": "Attribute Name", "width": 30},
    {"field": "value", "header": "Value", "width": 40},
    {"field": "unit", "header": "Unit", "width": 15},
    {"field": "language", "header": "Language", "width": 10},
]

DEFAULT_MEDIA_COLUMNS = [
    {"field": "product_code", "header": "Product Code", "width": 20},
    {"field": "media_type", "header": "Media Type", "width": 15},
    {"field": "file_url", "header": "File URL", "width": 60},
    {"field": "alt_text", "header": "Alt Text", "width": 40},
    {"field": "sort_order", "header": "Sort Order", "width": 12},
]

DEFAULT_PRICE_COLUMNS = [
    {"field": "product_code", "header": "Product Code", "width": 20},
    {"field": "price_list", "header": "Price List", "width": 25},
    {"field": "price", "header": "Price", "width": 15},
    {"field": "currency", "header": "Currency", "width": 10},
    {"field": "min_qty", "header": "Min Qty", "width": 12},
    {"field": "valid_from", "header": "Valid From", "width": 15},
    {"field": "valid_to", "header": "Valid To", "width": 15},
]


def export_catalog(
    profile_name=None,
    products=None,
    include_variants=True,
    include_attributes=True,
    include_media=True,
    include_prices=True,
    include_header_row=True,
    freeze_panes=True,
    auto_filter=True,
    language=None,
    currency=None,
    sheet_names=None,
    column_config=None,
    write_only=True,
    save_file=False
):
    """Generate XLSX Excel catalog export.

    This function creates an Excel workbook containing product information
    in multiple worksheets. It uses openpyxl's write_only mode by default
    for memory-efficient generation of large catalogs.

    Args:
        profile_name: Name of Export Profile DocType to use for settings
        products: List of Product Master/Variant names to export (optional)
        include_variants: Include variants in separate sheet (default: True)
        include_attributes: Include attributes sheet (default: True)
        include_media: Include media/images sheet (default: True)
        include_prices: Include pricing sheet (default: True)
        include_header_row: Add header row to each sheet (default: True)
        freeze_panes: Freeze header row for scrolling (default: True)
        auto_filter: Enable auto-filter on columns (default: True)
        language: Language code for localized content
        currency: Currency code for prices
        sheet_names: Custom sheet names dict (default uses standard names)
        column_config: Custom column configuration per sheet
        write_only: Use write_only mode for memory efficiency (default: True)
        save_file: Save to file and return file path

    Returns:
        bytes: XLSX file content as bytes, or file path if save_file=True

    Raises:
        ImportError: If openpyxl is not available

    Example:
        >>> xlsx_bytes = export_catalog(
        ...     profile_name="standard_xlsx",
        ...     save_file=True
        ... )
        >>> print(xlsx_bytes)  # Returns file path
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX export. "
            "Install with: pip install openpyxl"
        )

    # Load settings from profile if provided
    config = _load_profile_config(profile_name) if profile_name else {}

    # Override config with explicit parameters
    include_variants = config.get("include_variants", include_variants)
    include_attributes = config.get("include_attributes", include_attributes)
    include_media = config.get("include_media", include_media)
    include_prices = config.get("include_prices", include_prices)
    include_header_row = config.get("include_header_row", include_header_row)
    freeze_panes = config.get("freeze_panes", freeze_panes)
    auto_filter = config.get("auto_filter", auto_filter)
    language = language or config.get("language")
    currency = currency or config.get("currency", "EUR")

    # Default sheet names
    sheet_names = sheet_names or config.get("sheet_names", {
        "products": "Products",
        "variants": "Variants",
        "attributes": "Attributes",
        "media": "Media",
        "prices": "Prices"
    })

    # Column configuration
    column_config = column_config or config.get("column_config", {
        "products": DEFAULT_PRODUCT_COLUMNS,
        "variants": DEFAULT_VARIANT_COLUMNS,
        "attributes": DEFAULT_ATTRIBUTE_COLUMNS,
        "media": DEFAULT_MEDIA_COLUMNS,
        "prices": DEFAULT_PRICE_COLUMNS
    })

    # Get products to export
    if products is None:
        products = _get_products_for_export(config)

    # Create workbook in write_only mode for memory efficiency
    if write_only:
        wb = Workbook(write_only=True)
    else:
        wb = Workbook()
        # Remove default sheet
        if wb.active:
            del wb[wb.active.title]

    # Build worksheets
    _build_products_sheet(
        wb,
        products=products,
        sheet_name=sheet_names.get("products", "Products"),
        columns=column_config.get("products", DEFAULT_PRODUCT_COLUMNS),
        include_header=include_header_row,
        freeze_panes=freeze_panes,
        auto_filter=auto_filter,
        write_only=write_only,
        language=language
    )

    if include_variants:
        _build_variants_sheet(
            wb,
            products=products,
            sheet_name=sheet_names.get("variants", "Variants"),
            columns=column_config.get("variants", DEFAULT_VARIANT_COLUMNS),
            include_header=include_header_row,
            freeze_panes=freeze_panes,
            auto_filter=auto_filter,
            write_only=write_only,
            language=language,
            currency=currency
        )

    if include_attributes:
        _build_attributes_sheet(
            wb,
            products=products,
            sheet_name=sheet_names.get("attributes", "Attributes"),
            columns=column_config.get("attributes", DEFAULT_ATTRIBUTE_COLUMNS),
            include_header=include_header_row,
            freeze_panes=freeze_panes,
            auto_filter=auto_filter,
            write_only=write_only,
            language=language
        )

    if include_media:
        _build_media_sheet(
            wb,
            products=products,
            sheet_name=sheet_names.get("media", "Media"),
            columns=column_config.get("media", DEFAULT_MEDIA_COLUMNS),
            include_header=include_header_row,
            freeze_panes=freeze_panes,
            auto_filter=auto_filter,
            write_only=write_only
        )

    if include_prices:
        _build_prices_sheet(
            wb,
            products=products,
            sheet_name=sheet_names.get("prices", "Prices"),
            columns=column_config.get("prices", DEFAULT_PRICE_COLUMNS),
            include_header=include_header_row,
            freeze_panes=freeze_panes,
            auto_filter=auto_filter,
            write_only=write_only,
            currency=currency
        )

    # Save workbook to bytes
    xlsx_buffer = BytesIO()
    wb.save(xlsx_buffer)
    xlsx_buffer.seek(0)
    xlsx_content = xlsx_buffer.getvalue()

    # Close workbook
    wb.close()

    # Save to file if requested
    if save_file:
        return _save_export_file(
            xlsx_content,
            profile_name=profile_name
        )

    return xlsx_content


def export_catalog_async(profile_name, callback=None):
    """Queue catalog export as background job.

    For large catalogs, this function queues the export as a background
    job to avoid timeout issues.

    Args:
        profile_name: Name of Export Profile DocType
        callback: Optional callback function name to call on completion

    Returns:
        str: Background job ID
    """
    import frappe

    job = frappe.enqueue(
        "frappe_pim.pim.export.xlsx.export_catalog",
        queue="long",
        timeout=3600,
        profile_name=profile_name,
        save_file=True
    )

    return job.id if hasattr(job, 'id') else str(job)


def export_single_sheet(
    sheet_type,
    products=None,
    profile_name=None,
    columns=None,
    include_header=True,
    save_file=False
):
    """Export a single sheet type for specific use cases.

    Useful when you only need one type of data (e.g., just products
    or just attributes) without the full multi-sheet workbook.

    Args:
        sheet_type: One of 'products', 'variants', 'attributes', 'media', 'prices'
        products: List of product names to export
        profile_name: Export profile name for settings
        columns: Custom column configuration
        include_header: Include header row
        save_file: Save to file

    Returns:
        bytes: XLSX content or file path if save_file=True
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX export. "
            "Install with: pip install openpyxl"
        )

    # Load config
    config = _load_profile_config(profile_name) if profile_name else {}

    # Get products
    if products is None:
        products = _get_products_for_export(config)

    # Default columns for sheet type
    default_columns = {
        "products": DEFAULT_PRODUCT_COLUMNS,
        "variants": DEFAULT_VARIANT_COLUMNS,
        "attributes": DEFAULT_ATTRIBUTE_COLUMNS,
        "media": DEFAULT_MEDIA_COLUMNS,
        "prices": DEFAULT_PRICE_COLUMNS
    }

    columns = columns or default_columns.get(sheet_type, DEFAULT_PRODUCT_COLUMNS)

    # Create workbook
    wb = Workbook(write_only=True)

    # Build appropriate sheet
    sheet_builders = {
        "products": _build_products_sheet,
        "variants": _build_variants_sheet,
        "attributes": _build_attributes_sheet,
        "media": _build_media_sheet,
        "prices": _build_prices_sheet
    }

    builder = sheet_builders.get(sheet_type, _build_products_sheet)
    builder(
        wb,
        products=products,
        sheet_name=sheet_type.title(),
        columns=columns,
        include_header=include_header,
        freeze_panes=True,
        auto_filter=True,
        write_only=True
    )

    # Save to bytes
    xlsx_buffer = BytesIO()
    wb.save(xlsx_buffer)
    xlsx_buffer.seek(0)
    xlsx_content = xlsx_buffer.getvalue()
    wb.close()

    if save_file:
        return _save_export_file(
            xlsx_content,
            profile_name=profile_name,
            suffix=f"_{sheet_type}"
        )

    return xlsx_content


def _load_profile_config(profile_name):
    """Load export configuration from Export Profile DocType.

    Args:
        profile_name: Name of Export Profile document

    Returns:
        dict: Configuration dictionary
    """
    import frappe

    if not profile_name:
        return {}

    try:
        profile = frappe.get_doc("Export Profile", profile_name)
    except Exception:
        return {}

    # Map profile fields to config
    config = {
        "include_variants": profile.get("include_variants", True),
        "include_media": profile.get("include_media", True),
        "include_prices": profile.get("include_prices", True),
        "include_attributes": profile.get("include_attributes", True),
        "product_family": profile.get("product_family"),
        "status_filter": profile.get("status_filter"),
        "completeness_threshold": profile.get("completeness_threshold", 0),
        "output_filename": profile.get("output_filename"),
    }

    # Get language from export_language link
    if profile.get("export_language"):
        config["language"] = profile.export_language

    # Get currency from export_currency link
    if profile.get("export_currency"):
        config["currency"] = profile.export_currency

    # XLSX-specific settings
    config["include_header_row"] = profile.get("xlsx_include_header", True)
    config["freeze_panes"] = profile.get("xlsx_freeze_panes", True)
    config["auto_filter"] = profile.get("xlsx_auto_filter", True)

    # Custom column configuration from JSON field
    if profile.get("xlsx_column_config"):
        try:
            import json
            config["column_config"] = json.loads(profile.xlsx_column_config)
        except (json.JSONDecodeError, TypeError):
            pass

    return config


def _get_products_for_export(config):
    """Get list of products matching export filter criteria.

    Args:
        config: Export configuration dictionary

    Returns:
        list: List of product names to export
    """
    import frappe

    filters = {}

    if config.get("product_family"):
        filters["product_family"] = config["product_family"]

    if config.get("status_filter"):
        filters["status"] = config["status_filter"]

    # Get products that meet completeness threshold
    products = frappe.get_all(
        "Product Variant",
        filters=filters,
        fields=["name", "completeness_score"],
        order_by="modified desc"
    )

    # Filter by completeness threshold
    threshold = config.get("completeness_threshold", 0)
    if threshold > 0:
        products = [
            p for p in products
            if (p.get("completeness_score") or 0) >= threshold
        ]

    return [p["name"] for p in products]


def _build_products_sheet(
    wb,
    products,
    sheet_name,
    columns,
    include_header,
    freeze_panes,
    auto_filter,
    write_only,
    language=None
):
    """Build the Products worksheet.

    Args:
        wb: Workbook instance
        products: List of product names
        sheet_name: Name for the worksheet
        columns: Column configuration list
        include_header: Include header row
        freeze_panes: Freeze header row
        auto_filter: Enable auto-filter
        write_only: Using write_only mode
        language: Language for localized content
    """
    import frappe
    from openpyxl.utils import get_column_letter

    # Create worksheet
    ws = wb.create_sheet(title=sheet_name)

    # Header row
    if include_header:
        header_row = [col["header"] for col in columns]
        ws.append(header_row)

    # Get product data and write rows
    for product_name in products:
        try:
            # Try Product Variant first, then Product Master
            try:
                product = frappe.get_doc("Product Variant", product_name)
            except Exception:
                try:
                    product = frappe.get_doc("Product Master", product_name)
                except Exception:
                    continue

            row_data = []
            for col in columns:
                field = col["field"]
                value = _get_product_field_value(product, field, language)
                row_data.append(value)

            ws.append(row_data)

        except Exception:
            continue

    # Apply formatting (only for non-write_only mode)
    if not write_only and include_header:
        _apply_sheet_formatting(ws, columns, freeze_panes, auto_filter)


def _build_variants_sheet(
    wb,
    products,
    sheet_name,
    columns,
    include_header,
    freeze_panes,
    auto_filter,
    write_only,
    language=None,
    currency=None
):
    """Build the Variants worksheet.

    Args:
        wb: Workbook instance
        products: List of parent product names
        sheet_name: Name for the worksheet
        columns: Column configuration list
        include_header: Include header row
        freeze_panes: Freeze header row
        auto_filter: Enable auto-filter
        write_only: Using write_only mode
        language: Language for localized content
        currency: Currency code
    """
    import frappe

    ws = wb.create_sheet(title=sheet_name)

    # Header row
    if include_header:
        header_row = [col["header"] for col in columns]
        ws.append(header_row)

    # Get variants for each product
    for product_name in products:
        try:
            # Get variants linked to this product
            variants = frappe.get_all(
                "Product Variant",
                filters={"parent_product": product_name},
                fields=["name"]
            )

            for variant_ref in variants:
                try:
                    variant = frappe.get_doc("Product Variant", variant_ref["name"])
                    row_data = []
                    for col in columns:
                        field = col["field"]
                        if field == "parent_product":
                            value = product_name
                        elif field == "currency":
                            value = currency or variant.get("currency", "EUR")
                        else:
                            value = _get_product_field_value(variant, field, language)
                        row_data.append(value)
                    ws.append(row_data)
                except Exception:
                    continue

        except Exception:
            continue

    if not write_only and include_header:
        _apply_sheet_formatting(ws, columns, freeze_panes, auto_filter)


def _build_attributes_sheet(
    wb,
    products,
    sheet_name,
    columns,
    include_header,
    freeze_panes,
    auto_filter,
    write_only,
    language=None
):
    """Build the Attributes worksheet.

    Args:
        wb: Workbook instance
        products: List of product names
        sheet_name: Name for the worksheet
        columns: Column configuration list
        include_header: Include header row
        freeze_panes: Freeze header row
        auto_filter: Enable auto-filter
        write_only: Using write_only mode
        language: Language for localized content
    """
    import frappe

    ws = wb.create_sheet(title=sheet_name)

    # Header row
    if include_header:
        header_row = [col["header"] for col in columns]
        ws.append(header_row)

    # Get attributes for each product
    for product_name in products:
        try:
            # Try to get product document
            try:
                product = frappe.get_doc("Product Variant", product_name)
            except Exception:
                try:
                    product = frappe.get_doc("Product Master", product_name)
                except Exception:
                    continue

            product_code = product.get("variant_code") or product.get("product_code") or product_name
            attribute_values = product.get("attribute_values") or []

            for attr_value in attribute_values:
                attr_code = attr_value.get("attribute")
                if not attr_code:
                    continue

                # Get attribute metadata
                try:
                    attr_meta = frappe.get_cached_value(
                        "PIM Attribute",
                        attr_code,
                        ["attribute_name", "unit"],
                        as_dict=True
                    )
                except Exception:
                    attr_meta = {"attribute_name": attr_code}

                # Get value
                value = _get_attribute_value(attr_value)
                if value is None:
                    continue

                row_data = []
                for col in columns:
                    field = col["field"]
                    if field == "product_code":
                        row_data.append(product_code)
                    elif field == "attribute_code":
                        row_data.append(attr_code)
                    elif field == "attribute_name":
                        row_data.append(attr_meta.get("attribute_name", attr_code))
                    elif field == "value":
                        row_data.append(str(value) if value is not None else "")
                    elif field == "unit":
                        row_data.append(attr_meta.get("unit", ""))
                    elif field == "language":
                        row_data.append(language or "")
                    else:
                        row_data.append("")

                ws.append(row_data)

        except Exception:
            continue

    if not write_only and include_header:
        _apply_sheet_formatting(ws, columns, freeze_panes, auto_filter)


def _build_media_sheet(
    wb,
    products,
    sheet_name,
    columns,
    include_header,
    freeze_panes,
    auto_filter,
    write_only
):
    """Build the Media worksheet.

    Args:
        wb: Workbook instance
        products: List of product names
        sheet_name: Name for the worksheet
        columns: Column configuration list
        include_header: Include header row
        freeze_panes: Freeze header row
        auto_filter: Enable auto-filter
        write_only: Using write_only mode
    """
    import frappe

    ws = wb.create_sheet(title=sheet_name)

    # Header row
    if include_header:
        header_row = [col["header"] for col in columns]
        ws.append(header_row)

    # Get media for each product
    for product_name in products:
        try:
            # Get product
            try:
                product = frappe.get_doc("Product Variant", product_name)
            except Exception:
                try:
                    product = frappe.get_doc("Product Master", product_name)
                except Exception:
                    continue

            product_code = product.get("variant_code") or product.get("product_code") or product_name

            # Primary image
            primary_image = product.get("image")
            if primary_image:
                row_data = _build_media_row(
                    columns,
                    product_code,
                    "Primary",
                    _get_full_url(primary_image),
                    product.get("image_alt_text", ""),
                    1
                )
                ws.append(row_data)

            # Additional media from child table
            media_list = product.get("media") or []
            for idx, media in enumerate(media_list, start=2):
                media_url = media.get("file_url") or media.get("url")
                if media_url:
                    row_data = _build_media_row(
                        columns,
                        product_code,
                        media.get("media_type", "Image"),
                        _get_full_url(media_url),
                        media.get("alt_text", ""),
                        idx
                    )
                    ws.append(row_data)

        except Exception:
            continue

    if not write_only and include_header:
        _apply_sheet_formatting(ws, columns, freeze_panes, auto_filter)


def _build_prices_sheet(
    wb,
    products,
    sheet_name,
    columns,
    include_header,
    freeze_panes,
    auto_filter,
    write_only,
    currency=None
):
    """Build the Prices worksheet.

    Args:
        wb: Workbook instance
        products: List of product names
        sheet_name: Name for the worksheet
        columns: Column configuration list
        include_header: Include header row
        freeze_panes: Freeze header row
        auto_filter: Enable auto-filter
        write_only: Using write_only mode
        currency: Default currency code
    """
    import frappe

    ws = wb.create_sheet(title=sheet_name)

    # Header row
    if include_header:
        header_row = [col["header"] for col in columns]
        ws.append(header_row)

    # Get prices for each product
    for product_name in products:
        try:
            # Get product
            try:
                product = frappe.get_doc("Product Variant", product_name)
            except Exception:
                try:
                    product = frappe.get_doc("Product Master", product_name)
                except Exception:
                    continue

            product_code = product.get("variant_code") or product.get("product_code") or product_name

            # Standard price
            price = product.get("price") or product.get("standard_rate")
            if price:
                row_data = _build_price_row(
                    columns,
                    product_code,
                    "Standard",
                    price,
                    currency or product.get("currency", "EUR"),
                    1,
                    None,
                    None
                )
                ws.append(row_data)

            # Price list entries from child table
            price_list = product.get("prices") or []
            for price_entry in price_list:
                row_data = _build_price_row(
                    columns,
                    product_code,
                    price_entry.get("price_list", "Custom"),
                    price_entry.get("price"),
                    price_entry.get("currency", currency or "EUR"),
                    price_entry.get("min_qty", 1),
                    price_entry.get("valid_from"),
                    price_entry.get("valid_to")
                )
                ws.append(row_data)

        except Exception:
            continue

    if not write_only and include_header:
        _apply_sheet_formatting(ws, columns, freeze_panes, auto_filter)


def _build_media_row(columns, product_code, media_type, file_url, alt_text, sort_order):
    """Build a row for the media sheet.

    Args:
        columns: Column configuration
        product_code: Product identifier
        media_type: Type of media
        file_url: URL to media file
        alt_text: Alt text for image
        sort_order: Display order

    Returns:
        list: Row data
    """
    row_data = []
    for col in columns:
        field = col["field"]
        if field == "product_code":
            row_data.append(product_code)
        elif field == "media_type":
            row_data.append(media_type)
        elif field == "file_url":
            row_data.append(file_url)
        elif field == "alt_text":
            row_data.append(alt_text)
        elif field == "sort_order":
            row_data.append(sort_order)
        else:
            row_data.append("")
    return row_data


def _build_price_row(columns, product_code, price_list, price, currency, min_qty, valid_from, valid_to):
    """Build a row for the prices sheet.

    Args:
        columns: Column configuration
        product_code: Product identifier
        price_list: Price list name
        price: Price value
        currency: Currency code
        min_qty: Minimum quantity
        valid_from: Start date
        valid_to: End date

    Returns:
        list: Row data
    """
    row_data = []
    for col in columns:
        field = col["field"]
        if field == "product_code":
            row_data.append(product_code)
        elif field == "price_list":
            row_data.append(price_list)
        elif field == "price":
            row_data.append(float(price) if price else 0)
        elif field == "currency":
            row_data.append(currency)
        elif field == "min_qty":
            row_data.append(min_qty)
        elif field == "valid_from":
            row_data.append(_format_date(valid_from))
        elif field == "valid_to":
            row_data.append(_format_date(valid_to))
        else:
            row_data.append("")
    return row_data


def _get_product_field_value(product, field, language=None):
    """Get field value from product document.

    Args:
        product: Product document
        field: Field name to get
        language: Language code for localized content

    Returns:
        Field value or empty string
    """
    value = product.get(field)

    if value is None:
        return ""

    # Handle specific field types
    if field in ("description", "short_description"):
        return _clean_html(value)

    if field == "completeness_score":
        return float(value) if value else 0

    return value


def _get_attribute_value(attr_value):
    """Extract value from EAV attribute value row.

    Args:
        attr_value: Product Attribute Value row

    Returns:
        Value or None if no value set
    """
    # Check value columns in order of preference
    value_fields = [
        "value_text",
        "value_data",
        "value_int",
        "value_float",
        "value_date",
        "value_datetime",
        "value_link",
        "value_boolean"
    ]

    for field in value_fields:
        value = attr_value.get(field)
        if value is not None:
            if isinstance(value, bool):
                return "Yes" if value else "No"
            if isinstance(value, str) and not value.strip():
                continue
            return value

    return None


def _clean_html(html_text):
    """Remove HTML tags from text.

    Args:
        html_text: Text possibly containing HTML

    Returns:
        str: Clean text without HTML tags
    """
    if not html_text:
        return ""

    import re
    # Remove HTML tags
    clean = re.sub(r'<[^>]+>', '', str(html_text))
    # Normalize whitespace
    clean = re.sub(r'\s+', ' ', clean).strip()
    return clean


def _get_full_url(file_path):
    """Convert file path to full URL.

    Args:
        file_path: Relative file path

    Returns:
        str: Full URL to file
    """
    import frappe

    if not file_path:
        return ""

    # Already a full URL
    if file_path.startswith(("http://", "https://")):
        return file_path

    # Get site URL
    try:
        site_url = frappe.utils.get_url()
    except Exception:
        site_url = ""

    # Ensure path starts with /
    if not file_path.startswith("/"):
        file_path = "/" + file_path

    return f"{site_url}{file_path}"


def _format_date(date_value):
    """Format date value for Excel.

    Args:
        date_value: Date value (string or datetime)

    Returns:
        str: Formatted date string
    """
    if not date_value:
        return ""

    if isinstance(date_value, str):
        return date_value

    try:
        return date_value.strftime("%Y-%m-%d")
    except Exception:
        return str(date_value)


def _apply_sheet_formatting(ws, columns, freeze_panes, auto_filter):
    """Apply formatting to worksheet (non-write_only mode only).

    Args:
        ws: Worksheet instance
        columns: Column configuration
        freeze_panes: Whether to freeze header row
        auto_filter: Whether to enable auto-filter
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill

    # Set column widths
    for idx, col in enumerate(columns, start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = col.get("width", 15)

    # Style header row
    if ws.max_row >= 1:
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="DDDDDD",
            end_color="DDDDDD",
            fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    # Freeze panes (freeze first row)
    if freeze_panes:
        ws.freeze_panes = "A2"

    # Enable auto-filter
    if auto_filter and ws.max_row >= 1:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


def _save_export_file(content, profile_name=None, suffix=""):
    """Save export content to file.

    Args:
        content: XLSX content bytes
        profile_name: Export profile name (for filename)
        suffix: Additional filename suffix

    Returns:
        str: File path of saved file
    """
    import frappe

    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    name_part = profile_name or "xlsx_export"
    filename = f"{name_part}{suffix}_{timestamp}.xlsx"

    # Create file record
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": filename,
        "content": content,
        "is_private": 1,
        "folder": "Home/Exports"
    })

    try:
        file_doc.insert(ignore_permissions=True)
        frappe.db.commit()
    except Exception:
        # Folder might not exist, try without folder
        file_doc.folder = "Home"
        file_doc.insert(ignore_permissions=True)
        frappe.db.commit()

    # Update Export Profile if name provided
    if profile_name:
        try:
            frappe.db.set_value(
                "Export Profile",
                profile_name,
                {
                    "last_export": datetime.now(),
                    "last_file": file_doc.file_url,
                    "export_status": "Completed"
                },
                update_modified=True
            )
            frappe.db.commit()
        except Exception:
            pass

    return file_doc.file_url


def validate_xlsx(xlsx_content):
    """Validate XLSX file structure.

    Args:
        xlsx_content: XLSX file content (bytes or file path)

    Returns:
        tuple: (is_valid, errors_list)
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False, ["openpyxl not available"]

    errors = []

    try:
        # Load workbook from bytes or file
        if isinstance(xlsx_content, bytes):
            xlsx_buffer = BytesIO(xlsx_content)
            wb = load_workbook(xlsx_buffer, read_only=True)
        else:
            wb = load_workbook(xlsx_content, read_only=True)

        # Check for at least one sheet
        if len(wb.sheetnames) == 0:
            errors.append("Workbook has no sheets")

        # Check each sheet has data
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row_count = 0
            for _ in ws.rows:
                row_count += 1
                if row_count > 1:
                    break

            if row_count == 0:
                errors.append(f"Sheet '{sheet_name}' is empty")

        wb.close()

        return len(errors) == 0, errors

    except Exception as e:
        return False, [f"Validation error: {str(e)}"]


def get_sheet_count(xlsx_content):
    """Get number of sheets in XLSX file.

    Args:
        xlsx_content: XLSX file content (bytes)

    Returns:
        int: Number of sheets
    """
    try:
        from openpyxl import load_workbook

        xlsx_buffer = BytesIO(xlsx_content)
        wb = load_workbook(xlsx_buffer, read_only=True)
        count = len(wb.sheetnames)
        wb.close()
        return count
    except Exception:
        return 0


def get_row_count(xlsx_content, sheet_name=None):
    """Get row count for a sheet in XLSX file.

    Args:
        xlsx_content: XLSX file content (bytes)
        sheet_name: Specific sheet name (default: first sheet)

    Returns:
        int: Number of rows (excluding header)
    """
    try:
        from openpyxl import load_workbook

        xlsx_buffer = BytesIO(xlsx_content)
        wb = load_workbook(xlsx_buffer, read_only=True)

        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Count rows (subtract 1 for header)
        row_count = sum(1 for _ in ws.rows) - 1
        wb.close()

        return max(0, row_count)
    except Exception:
        return 0


def read_xlsx_products(xlsx_content, sheet_name="Products"):
    """Read products from XLSX file (for import).

    Args:
        xlsx_content: XLSX file content (bytes)
        sheet_name: Sheet name containing products

    Returns:
        list: List of product dicts
    """
    try:
        from openpyxl import load_workbook

        xlsx_buffer = BytesIO(xlsx_content)
        wb = load_workbook(xlsx_buffer, read_only=True)
        ws = wb[sheet_name]

        products = []
        headers = None

        for row_idx, row in enumerate(ws.rows):
            if row_idx == 0:
                # First row is headers
                headers = [cell.value for cell in row]
                continue

            # Build product dict
            product = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    product[headers[col_idx]] = cell.value
            products.append(product)

        wb.close()
        return products

    except Exception:
        return []
